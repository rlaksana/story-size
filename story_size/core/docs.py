from pathlib import Path
import pypdf
import docx
import openpyxl

def read_pdf(file_path: Path) -> str:
    """Reads text from a PDF file."""
    reader = pypdf.PdfReader(file_path)
    text = []
    for page in reader.pages:
        text.append(page.extract_text())
    return "\n".join(text)

def read_docx(file_path: Path) -> str:
    """Reads text from a DOCX file."""
    doc = docx.Document(file_path)
    text = []
    for para in doc.paragraphs:
        text.append(para.text)
    return "\n".join(text)

def read_xlsx(file_path: Path, max_rows: int = 100) -> str:
    """Reads text from an XLSX file with row limit to prevent huge content."""
    workbook = openpyxl.load_workbook(file_path)
    text = []

    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        row_count = 0

        for row in sheet.iter_rows():
            if row_count >= max_rows:
                text.append(f"... (truncated after {max_rows} rows in sheet '{sheetname}')")
                break

            row_text = []
            for cell in row:
                if cell.value:
                    row_text.append(str(cell.value))
            if row_text:  # Only add non-empty rows
                text.append(" ".join(row_text))
            row_count += 1

    return "\n".join(text)

def read_documents(docs_dir: Path) -> str:
    """
    Reads all .md, .txt, .pdf, .docx, and .xlsx files from the given directory,
    concatenates their content, and returns it as a single string.
    """
    content = []
    for file_path in docs_dir.rglob("*.md"):
        try:
            content.append(file_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    for file_path in docs_dir.rglob("*.txt"):
        try:
            content.append(file_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    for file_path in docs_dir.rglob("*.pdf"):
        try:
            content.append(read_pdf(file_path))
        except Exception:
            pass
    for file_path in docs_dir.rglob("*.docx"):
        try:
            content.append(read_docx(file_path))
        except Exception:
            pass
    for file_path in docs_dir.rglob("*.xlsx"):
        try:
            content.append(read_xlsx(file_path))
        except Exception:
            pass
    return "\n\n".join(content)

def read_documents_limited(docs_dir: Path, max_content_length: int = 50000) -> str:
    """
    Reads all .md, .txt, .pdf, .docx, and .xlsx files from the given directory,
    concatenates their content, and returns it as a single string with size limits.

    Args:
        docs_dir: Directory containing document files
        max_content_length: Maximum total characters to prevent API limits
    """
    content = []
    current_length = 0

    # Read markdown files first (usually most important)
    for file_path in docs_dir.rglob("*.md"):
        try:
            file_content = file_path.read_text(encoding="utf-8")
            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:  # Only add if we have meaningful space
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
        except Exception:
            pass

    # Read text files
    for file_path in docs_dir.rglob("*.txt"):
        try:
            file_content = file_path.read_text(encoding="utf-8")
            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
        except Exception:
            pass

    # Read PDF files with limit
    for file_path in docs_dir.rglob("*.pdf"):
        try:
            file_content = read_pdf_limited(file_path, max_chars=10000)
            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
        except Exception:
            pass

    # Read DOCX files with limit
    for file_path in docs_dir.rglob("*.docx"):
        try:
            file_content = read_docx(file_path)
            if len(file_content) > 15000:  # Limit DOCX content
                file_content = file_content[:15000] + "\n... (truncated)"
            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
        except Exception:
            pass

    # Read Excel files with row limit (most restrictive)
    for file_path in docs_dir.rglob("*.xlsx"):
        try:
            file_content = read_xlsx(file_path, max_rows=20)  # Very limited Excel rows
            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
        except Exception:
            pass

    return "\n\n".join(content)

def read_pdf_limited(file_path: Path, max_chars: int = 10000) -> str:
    """Reads limited text from a PDF file."""
    reader = pypdf.PdfReader(file_path)
    text = []
    current_chars = 0

    for page in reader.pages:
        page_text = page.extract_text()
        if current_chars + len(page_text) > max_chars:
            remaining_chars = max_chars - current_chars
            if remaining_chars > 0:
                text.append(page_text[:remaining_chars] + "... (truncated)")
            break
        text.append(page_text)
        current_chars += len(page_text)

    return "\n".join(text)