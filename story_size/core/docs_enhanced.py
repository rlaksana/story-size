"""
Enhanced document reader with image support.
This module provides image extraction and OCR capabilities for story size estimation.
"""

from pathlib import Path
from typing import Dict, Any
import logging

import pypdf
import docx
import openpyxl
from .image_processing import ImageProcessor

logger = logging.getLogger(__name__)


def read_documents_with_images(docs_dir: Path, max_content_length: int = 500000) -> Dict[str, Any]:
    """
    Enhanced document reader that extracts both text and image information.

    Args:
        docs_dir: Directory containing document files
        max_content_length: Maximum total characters (500K ≈ 125K tokens) to prevent API limits

    Returns:
        Dictionary containing:
        - text_content: Concatenated text from all files
        - image_analysis: List of image analysis results
        - total_images: Total number of images found
        - complexity_indicators: Summary of visual elements found
    """
    content = []
    image_analysis = []
    current_length = 0

    processor = ImageProcessor()

    logger.info(f"Starting enhanced document processing for: {docs_dir}")

    # Read markdown files first (usually most important)
    for file_path in docs_dir.rglob("*.md"):
        if current_length >= max_content_length:
            break

        try:
            file_content = file_path.read_text(encoding="utf-8")
            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
            logger.debug(f"Read markdown: {file_path.name}")
        except Exception as e:
            logger.warning(f"Failed to read {file_path}: {e}")
            pass

    # Read text files
    for file_path in docs_dir.rglob("*.txt"):
        if current_length >= max_content_length:
            break

        try:
            file_content = file_path.read_text(encoding="utf-8")
            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
            logger.debug(f"Read text: {file_path.name}")
        except Exception as e:
            logger.warning(f"Failed to read {file_path}: {e}")
            pass

    # Read PDF files with image extraction
    for file_path in docs_dir.rglob("*.pdf"):
        if current_length >= max_content_length:
            break

        try:
            # Extract text
            file_content = read_pdf_limited(file_path, max_chars=100000)

            # Extract and analyze images
            logger.info(f"Extracting images from PDF: {file_path.name}")
            images = processor.extract_pdf_images(file_path)

            for img_info in images:
                analysis = processor.analyze_image_for_story_estimation(img_info['image'])
                image_data = {
                    'file': str(file_path),
                    'file_name': file_path.name,
                    'page': img_info.get('page', 0),
                    'index': img_info.get('index', 0),
                    'size': img_info.get('size', (0, 0)),
                    'analysis': analysis
                }
                image_analysis.append(image_data)

                # Add OCR text to content if available and meaningful
                if analysis.get('has_text') and analysis.get('text_length', 0) > 50:
                    page_num = img_info.get('page', 0)
                    ocr_text = f"\n\n[Extracted text from image on page {page_num}]:\n{analysis.get('full_text', '')}"

                    if current_length + len(ocr_text) < max_content_length:
                        file_content += ocr_text
                        current_length += len(ocr_text)
                        logger.debug(f"Added OCR text from page {page_num}")

            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
            logger.info(f"Processed PDF: {file_path.name} with {len(images)} images")

        except Exception as e:
            logger.error(f"Failed to process PDF {file_path}: {e}")
            pass

    # Read DOCX files with image extraction
    for file_path in docs_dir.rglob("*.docx"):
        if current_length >= max_content_length:
            break

        try:
            file_content = read_docx(file_path)

            # Limit DOCX content
            if len(file_content) > 100000:
                file_content = file_content[:100000] + "\n... (truncated)"

            # Extract and analyze images
            logger.info(f"Extracting images from DOCX: {file_path.name}")
            images = processor.extract_docx_images(file_path)

            for img_info in images:
                analysis = processor.analyze_image_for_story_estimation(img_info['image'])
                image_data = {
                    'file': str(file_path),
                    'file_name': file_path.name,
                    'filename': img_info.get('filename', ''),
                    'size': img_info.get('size', (0, 0)),
                    'analysis': analysis
                }
                image_analysis.append(image_data)

                # Add OCR text to content if available and meaningful
                if analysis.get('has_text') and analysis.get('text_length', 0) > 50:
                    ocr_text = f"\n\n[Extracted text from image in {file_path.name}]:\n{analysis.get('full_text', '')}"

                    if current_length + len(ocr_text) < max_content_length:
                        file_content += ocr_text
                        current_length += len(ocr_text)
                        logger.debug(f"Added OCR text from DOCX image")

            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
            logger.info(f"Processed DOCX: {file_path.name} with {len(images)} images")

        except Exception as e:
            logger.error(f"Failed to process DOCX {file_path}: {e}")
            pass

    # Read Excel files (no image support for now)
    for file_path in docs_dir.rglob("*.xlsx"):
        if current_length >= max_content_length:
            break

        try:
            file_content = read_xlsx(file_path, max_rows=20)
            if current_length + len(file_content) > max_content_length:
                remaining_space = max_content_length - current_length
                if remaining_space > 100:
                    content.append(file_content[:remaining_space] + "\n... (truncated)")
                break
            content.append(file_content)
            current_length += len(file_content)
            logger.debug(f"Read Excel: {file_path.name}")
        except Exception as e:
            logger.warning(f"Failed to read {file_path}: {e}")
            pass

    # Calculate complexity indicators
    complexity_indicators = {
        'has_diagrams': any(img['analysis']['complexity_indicators'].get('has_diagram', False)
                          for img in image_analysis),
        'has_tables': any(img['analysis']['complexity_indicators'].get('has_table', False)
                        for img in image_analysis),
        'has_screenshots': any(img['analysis']['complexity_indicators'].get('has_screenshot', False)
                            for img in image_analysis),
        'has_forms': any(img['analysis']['complexity_indicators'].get('has_form', False)
                      for img in image_analysis),
        'has_workflows': any(img['analysis']['complexity_indicators'].get('has_workflow', False)
                          for img in image_analysis),
        'has_icons': any(img['analysis']['complexity_indicators'].get('has_icon', False)
                       for img in image_analysis)
    }

    # Calculate total complexity scores from images
    total_image_complexity = sum(
        img['analysis'].get('total_complexity_factor', 0)
        for img in image_analysis
    )

    result = {
        'text_content': "\n\n".join(content),
        'image_analysis': image_analysis,
        'total_images': len(image_analysis),
        'images_with_text': sum(1 for img in image_analysis if img['analysis'].get('has_text', False)),
        'total_ocr_chars': sum(img['analysis'].get('text_length', 0) for img in image_analysis),
        'complexity_indicators': complexity_indicators,
        'total_image_complexity': total_image_complexity,
        'files_processed': len(content)
    }

    logger.info(f"Document processing complete: {result['total_images']} images, "
               f"{result['images_with_text']} with text, {result['total_image_complexity']} complexity score")

    return result


def read_pdf_limited(file_path: Path, max_chars: int = 100000) -> str:
    """Reads text from a PDF file with character limit (default 100K)."""
    reader = pypdf.PdfReader(file_path)
    text = []
    total_chars = 0

    for page in reader.pages:
        page_text = page.extract_text()
        if total_chars + len(page_text) > max_chars:
            # Truncate to fit within limit
            remaining = max_chars - total_chars
            if remaining > 0:
                text.append(page_text[:remaining])
            break
        text.append(page_text)
        total_chars += len(page_text)

    return "\n".join(text)

def read_docx(file_path: Path) -> str:
    """Reads text from a DOCX file."""
    doc = docx.Document(file_path)
    text = []
    for para in doc.paragraphs:
        text.append(para.text)
    return "\n".join(text)

def read_xlsx(file_path: Path, max_rows: int = 100) -> str:
    """Reads text from an XLSX file with row limit to prevent huge content."""
    workbook = openpyxl.load_workbook(file_path)
    text = []

    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        row_count = 0

        for row in sheet.iter_rows():
            if row_count >= max_rows:
                text.append(f"... (truncated after {max_rows} rows in sheet '{sheetname}')")
                break

            row_text = []
            for cell in row:
                if cell.value:
                    row_text.append(str(cell.value))
            if row_text:  # Only add non-empty rows
                text.append(" ".join(row_text))
            row_count += 1

    return "\n".join(text)